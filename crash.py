
import hashlib
import hmac
import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Constants
BLOCK_HASH = "0000000000000000001b34dc6a1e86083f95500b096231436e9b25cbdd0075c4"
US_BLOCK_HASH = "000000000000000000066448f2f56069750fc40c718322766b6bdf63fdcf45b8"

def get_result(game_hash, use_us_block_hash=False):
    salt = US_BLOCK_HASH if use_us_block_hash else BLOCK_HASH
    
    hmac_obj = hmac.new(game_hash.encode('utf-8'), salt.encode('utf-8'), hashlib.sha256)
    hex_result = hmac_obj.hexdigest()[:8]
    dec_result = int(hex_result, 16)
    
    f = float((4294967296 / (dec_result + 1)) * (1 - 0.01))
    point = float(round(f * 100) / 100)
    
    return point

def get_prev_game(hash_code):
    m = hashlib.sha256()
    m.update(hash_code.encode("utf-8"))
    return m.hexdigest()

# Example usage
game_hash = input("Enter Starting Hash: ")
print("Enter Starting Hash:", game_hash)

first_game = "cc537daff38b7e3c6ed9c1254f215e7db1eb4303397ee53d6ee2b88364c0cb30"

results = []
count = 0

while game_hash != first_game:
    count += 1
    result = get_result(game_hash)
    results.append(result)
    game_hash = get_prev_game(game_hash)
    
    if count % 1000 == 0:
        print(f"Processed {count} games")

print(f"Total games processed: {count}")
print(f"First 10 results: {results[:10]}")
print(f"Last 10 results: {results[-10:]}")

# Save results to CSV file
csv_file_path = 'results.csv'
with open(csv_file_path, 'w', newline='') as csvfile:
    csv_writer = csv.writer(csvfile)
    for result in results:
        csv_writer.writerow([result])

print(f"Results have been saved to {csv_file_path}")

# Read the CSV file
csv_data = pd.read_csv('results.csv', header=None)

# Load the existing .xlsm workbook
workbook = load_workbook('crash.xlsm', keep_vba=True)

# Select the 'crash' sheet
sheet = workbook['crash']

# Overwrite data in column A
for r_idx, value in enumerate(csv_data.iloc[:, 0], 1):
    sheet.cell(row=r_idx, column=1, value=value)

# Get the last row with data in column A
last_row = sheet.max_row

# Extend columns B to AQ
for col in range(2, 43):  # 2 to 43 represents columns B to AQ
    current_col_length = len([cell for cell in sheet[sheet.cell(row=1, column=col).column_letter] if cell.value is not None])
    if current_col_length < last_row:
        for row in range(current_col_length + 1, last_row + 1):
            sheet.cell(row=row, column=col, value='')

# Save the workbook
workbook.save('crash.xlsm')

print("Excel file has been updated successfully.")